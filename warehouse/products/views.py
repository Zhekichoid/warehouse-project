from django.shortcuts import render
from openpyxl import load_workbook
from .models import Product
from django.http import HttpResponse
from django.views import generic


class IndexView(generic.ListView):
    template_name = "products/index.html"
    context_object_name = "product_list"

    def get_queryset(self):
        return Product.objects.all()[:5]
    
class DetailView(generic.DetailView):
    model = Product
    template_name = "products/detail.html"

    def get_queryset(self):
        return Product.objects.all()

def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            brand, product_type, product_id, initial_price, discount = row
            final_price = int(initial_price) * (100 - discount) / 100
            Product.objects.create(product_id=product_id, product_type=product_type, 
                                   brand=brand, initial_price=initial_price,
                                   discount=discount, final_price=final_price
                                   )

        return render(request, 'import_success.html')

    return render(request, 'import_form.html')
# Create your views here.
